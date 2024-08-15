from tkinter import *
import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
from tkinter import messagebox
import mysql.connector
import winreg
import openpyxl
from datetime import date
from fpdf import FPDF
import os

current = os.path.dirname(os.path.realpath(__file__))


def Mark_form_6_9(frame):
    form_6_9 = Frame(frame, height=651, width=1185)
    form_6_9.pack()

    back = PhotoImage(file=current+"/.ux/padma/.mark/.forms/.primary/ADD MARKS.png")
    Background = Label(form_6_9, image=back)
    Background.image = back
    Background.place(x=0, y=0, relheight=1, relwidth=1)

    fonts = ("Candara", 15)

    # Header
    try:
        def collect(text):
            try:
                key = winreg.OpenKey(winreg.HKEY_CURRENT_USER, r"SOFTWARE\PADMA")
                host = winreg.QueryValueEx(key, "Host")[0]
                user = winreg.QueryValueEx(key, "User")[0]
                password = winreg.QueryValueEx(key, "Pass")[0]
                winreg.CloseKey(key)

                con = mysql.connector.connect(host=host,
                                              user=user,
                                              password=password,
                                              database="pcc")
                mycursor = con.cursor(buffered=True)
            except:
                messagebox.showerror("Error", "Database Error")
                return

            query = "SELECT id, full_name FROM student WHERE full_name LIKE '%" + text + "%' " \
                                                                                         "OR id LIKE '%" + text + "%' ORDER BY id ASC"
            mycursor.execute(query)
            result = mycursor.fetchall()
            con.close()

            list.place(x=270, y=209)

            list.delete(0, END)
            for row in result:
                idd = str(row[0])
                list.insert(END, idd + " , " + row[1])

        def select():
            selected_suggestion = list.get(list.curselection()[0])
            selected_suggestion = selected_suggestion.split()
            Key = selected_suggestion[0]

            mark_index_entry.delete(0, END)
            mark_index_entry.insert(0, Key)

            list.place_forget()

        mark_index_entry = Entry(form_6_9, fg="#1b2d52", font=fonts)
        mark_index_entry.place(width=203, height=26, x=270, y=183)
        mark_index_entry.bind("<KeyRelease>", lambda event: collect(mark_index_entry.get()))

        Year = Entry(form_6_9, fg="#1b2d52", font=fonts)
        Year.place(width=150, height=25, x=532, y=183)

        options = ["1st", "2nd", "3rd"]
        Term = ttk.Combobox(form_6_9, values=options, font=fonts)
        Term.place(width=150, height=25, x=736, y=183)
        Term.set(options[0])

        options_g6 = ["6A", "6B", "6C", "6D", "6E", "6F"]
        options_g7 = ["7A", "7B", "7C", "7D", "7E", "7F"]
        options_g8 = ["8A", "6B", "8C", "8D", "8E", "8F"]
        options_g9 = ["9A", "9B", "9C", "9D", "9E", "9F"]
        options = options_g6 + options_g7 + options_g8 + options_g9
        Class = ttk.Combobox(form_6_9, values=options, font=fonts)
        Class.place(width=150, height=25, x=938, y=183)
        Class.set(options[0])
    except:
        pass

    # Sheet
    try:
        sinhala_mark_entry = Entry(form_6_9, fg="#1b2d52", font=fonts, bg="#e9edfa")
        sinhala_mark_entry.place(width=78, height=30, x=370, y=256)
        maths_mark_entry = Entry(form_6_9, fg="#1b2d52", font=fonts, bg="#e9edfa")
        maths_mark_entry.place(width=78, height=30, x=595, y=256)
        science_mark_entry = Entry(form_6_9, fg="#1b2d52", font=fonts, bg="#e9edfa")
        science_mark_entry.place(width=78, height=30, x=834, y=256)
        english_mark_entry = Entry(form_6_9, fg="#1b2d52", font=fonts, bg="#e9edfa")
        english_mark_entry.place(width=78, height=30, x=1057, y=256)
        religion_mark_entry = Entry(form_6_9, fg="#1b2d52", font=fonts, bg="#e9edfa")
        religion_mark_entry.place(width=78, height=30, x=370, y=291)
        history_mark_entry = Entry(form_6_9, fg="#1b2d52", font=fonts, bg="#e9edfa")
        history_mark_entry.place(width=78, height=30, x=595, y=291)
        geography_mark_entry = Entry(form_6_9, fg="#1b2d52", font=fonts, bg="#e9edfa")
        geography_mark_entry.place(width=78, height=30, x=834, y=291)
        ce_ls_mark_entry = Entry(form_6_9, fg="#1b2d52", font=fonts, bg="#e9edfa")
        ce_ls_mark_entry.place(width=78, height=30, x=1057, y=291)
        health_mark_entry = Entry(form_6_9, fg="#1b2d52", font=fonts, bg="#e9edfa")
        health_mark_entry.place(width=78, height=30, x=370, y=325)
        tamil_mark_entry = Entry(form_6_9, fg="#1b2d52", font=fonts, bg="#e9edfa")
        tamil_mark_entry.place(width=78, height=30, x=595, y=325)
        pts_mark_entry = Entry(form_6_9, fg="#1b2d52", font=fonts, bg="#e9edfa")
        pts_mark_entry.place(width=78, height=30, x=834, y=325)
        ict_mark_entry = Entry(form_6_9, fg="#1b2d52", font=fonts, bg="#e9edfa")
        ict_mark_entry.place(width=78, height=30, x=1057, y=325)

        art_mark_entry = Entry(form_6_9, fg="#1b2d52", font=fonts, bg="#e9edfa")
        art_mark_entry.place(width=78, height=30, x=370, y=370)
        dance_mark_entry = Entry(form_6_9, fg="#1b2d52", font=fonts, bg="#e9edfa")
        dance_mark_entry.place(width=78, height=30, x=595, y=370)
        music_mark_entry = Entry(form_6_9, fg="#1b2d52", font=fonts, bg="#e9edfa")
        music_mark_entry.place(width=78, height=30, x=834, y=370)
        drama_mark_entry = Entry(form_6_9, fg="#1b2d52", font=fonts, bg="#e9edfa")
        drama_mark_entry.place(width=78, height=30, x=1057, y=370)
    except:
        pass

    list = Listbox(form_6_9, height=10, width=35, bd=0)
    list.bind("<Double-Button-1>", lambda event: select())

    def show():
        if list.winfo_ismapped():
            list.place_forget()
        else:
            list.place(x=270, y=209)

    drop_btn = Button(form_6_9, text="⇅", bg="white", font=("Candara", 20, "bold"), bd=0, command=show)
    drop_btn.place(x=445, y=184, width=26, height=25)

    fonts = ("Candara", 14, "bold")

    def filler():
        if art_mark_entry.get() == "":
            art_mark_entry.insert(0, "#")
        if drama_mark_entry.get() == "":
            drama_mark_entry.insert(0, "#")
        if dance_mark_entry.get() == "":
            dance_mark_entry.insert(0, "#")
        if music_mark_entry.get() == "":
            music_mark_entry.insert(0, "#")
        if ict_mark_entry.get() == "":
            ict_mark_entry.insert(0, "#")
        add_mark()

    def add_mark():
        if mark_index_entry.get() == "" or \
                Year.get() == "" or \
                Term.get() == "" or \
                Class.get() == "":
            messagebox.showerror("Error", "Index , Year, Term, Class Required")

        elif religion_mark_entry.get() == "" or \
                sinhala_mark_entry.get() == "" or \
                english_mark_entry.get() == "" or \
                tamil_mark_entry.get() == "" or \
                science_mark_entry.get() == "" or \
                maths_mark_entry.get() == "" or \
                history_mark_entry.get() == "" or \
                health_mark_entry.get() == "" or \
                pts_mark_entry.get() == "" or \
                geography_mark_entry.get() == "" or \
                ce_ls_mark_entry.get() == "":
            messagebox.showerror("Error", "All Marks Required")
        elif art_mark_entry.get() == "" or dance_mark_entry.get() == "" or drama_mark_entry.get() == "" or \
                music_mark_entry.get() == "" or ict_mark_entry.get() == "":
            filler()
        else:
            try:
                key = winreg.OpenKey(winreg.HKEY_CURRENT_USER, r"SOFTWARE\PADMA")
                host = winreg.QueryValueEx(key, "Host")[0]
                user = winreg.QueryValueEx(key, "User")[0]
                password = winreg.QueryValueEx(key, "Pass")[0]
                winreg.CloseKey(key)

                con = mysql.connector.connect(host=host,
                                              user=user,
                                              password=password,
                                              database="pcc")
                mycursor = con.cursor(buffered=True)
            except:
                messagebox.showerror("Error", "Database Failed")
                return

            try:
                query = "create table exam(id integer,year integer,term varchar(5),class varchar(10)" \
                        ",exam_id integer PRIMARY KEY AUTO_INCREMENT)"
                mycursor.execute(query)
                query = "create table mark_uo(exam_id integer PRIMARY KEY AUTO_INCREMENT" \
                        ",sinhala integer" \
                        ",maths integer" \
                        ",science integer" \
                        ",english integer" \
                        ",religion integer" \
                        ",history integer" \
                        ",geograpy integer" \
                        ",ce integer" \
                        ",health integer" \
                        ",art varchar(5)" \
                        ",dance varchar(5)" \
                        ",music varchar(5)" \
                        ",drama varchar(5)" \
                        ",tamil integer" \
                        ",pts integer" \
                        ",ict varchar(5))"
                mycursor.execute(query)
            except:
                mycursor.execute("use pcc")

            query = 'select * from student where id=%s'
            parameters = [mark_index_entry.get()]
            mycursor.execute(query, parameters)
            result = mycursor.fetchone()

            if result is not None:
                query = "select * from exam where id=%s and year=%s and term=%s and class=%s"
                parameters = [mark_index_entry.get(), Year.get(), Term.get(), Class.get()]
                mycursor.execute(query, parameters)
                result = mycursor.fetchone()

                if result is not None:
                    messagebox.showerror("Error", "Data Already Available")
                    con.close()
                else:
                    try:
                        query = "insert into mark_uo " \
                                "(sinhala,maths,science,english,religion,history," \
                                "geograpy,ce,health,art,dance,music,drama ,tamil ,pts ,ict )" \
                                " values (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"
                        parameters = [sinhala_mark_entry.get(),
                                      maths_mark_entry.get(),
                                      science_mark_entry.get(),
                                      english_mark_entry.get(),
                                      religion_mark_entry.get(),
                                      history_mark_entry.get(),
                                      geography_mark_entry.get(),
                                      ce_ls_mark_entry.get(),
                                      health_mark_entry.get(),
                                      art_mark_entry.get(),
                                      dance_mark_entry.get(),
                                      music_mark_entry.get(),
                                      drama_mark_entry.get(),
                                      tamil_mark_entry.get(),
                                      pts_mark_entry.get(),
                                      ict_mark_entry.get()]
                        mycursor.execute(query, parameters)
                        query = "insert into exam(id,year,term,class) values (%s,%s,%s,%s)"
                        parameters = [mark_index_entry.get(), Year.get()
                                      , Term.get(), Class.get()]
                        mycursor.execute(query, parameters)
                        con.commit()
                        con.close()
                        messagebox.showinfo("Success", "DATA SUBMITTED SUCCESSFULLY")
                    except:
                        messagebox.showerror("Error", "SUBMIT FAILED")
                        con.close()
            else:
                messagebox.showerror("Error", "Invalid index or Unknown Student")
            con.close()
            print("pass")

    submit_marks_button = Button(form_6_9, text="SUBMIT", bg="#ce4912", fg="white", font=fonts, command=add_mark)
    submit_marks_button.place(x=253, y=416, width=172, height=34)

    def clear():
        sinhala_mark_entry.delete(0, END)
        maths_mark_entry.delete(0, END)
        science_mark_entry.delete(0, END)
        english_mark_entry.delete(0, END)
        religion_mark_entry.delete(0, END)
        history_mark_entry.delete(0, END)
        geography_mark_entry.delete(0, END)
        ce_ls_mark_entry.delete(0, END)
        health_mark_entry.delete(0, END)
        art_mark_entry.delete(0, END)
        dance_mark_entry.delete(0, END)
        music_mark_entry.delete(0, END)
        drama_mark_entry.delete(0, END)
        tamil_mark_entry.delete(0, END)
        pts_mark_entry.delete(0, END)
        ict_mark_entry.delete(0, END)

    clear_marks_button = Button(form_6_9, text="CLEAR", bg="#3b3086", fg="white", font=fonts, command=clear)
    clear_marks_button.place(x=445, y=416, width=172, height=34)

    EDIT = Button(form_6_9, text="EDIT", bg="#3b3086", fg="white", font=fonts
                  , command=lambda: (form_6_9.destroy(), edit(frame)))
    EDIT.place(x=22, y=295, width=172, height=34)

    VIEWER = Button(form_6_9, text="VIEWER", bg="#ce4912", fg="white", font=fonts
                    , command=lambda: (form_6_9.destroy(), Viewer(frame)))
    VIEWER.place(x=22, y=340, width=172, height=34)


def edit(frame):
    form_6_9_edit = Frame(frame, height=651, width=1185)
    form_6_9_edit.pack()

    back = PhotoImage(file=current+"/.ux/padma/.mark/.edit_forms/.primary/EDIT MARKS.png")
    Background = Label(form_6_9_edit, image=back)
    Background.image = back
    Background.place(x=0, y=0, relheight=1, relwidth=1)

    fonts = ("Candara", 15)

    try:
        def collect(text):
            try:
                key = winreg.OpenKey(winreg.HKEY_CURRENT_USER, r"SOFTWARE\PADMA")
                host = winreg.QueryValueEx(key, "Host")[0]
                user = winreg.QueryValueEx(key, "User")[0]
                password = winreg.QueryValueEx(key, "Pass")[0]
                winreg.CloseKey(key)

                con = mysql.connector.connect(host=host,
                                              user=user,
                                              password=password,
                                              database="pcc")
                mycursor = con.cursor(buffered=True)
            except:
                messagebox.showerror("Error", "Database Error")
                return

            query = "SELECT id, full_name FROM student WHERE full_name LIKE '%" + text + "%' " \
                                                                                         "OR id LIKE '%" + text + "%' ORDER BY id ASC"
            mycursor.execute(query)
            result = mycursor.fetchall()
            con.close()

            list.place(x=270, y=209)

            list.delete(0, END)
            for row in result:
                idd = str(row[0])
                list.insert(END, idd + " , " + row[1])

        def select():
            selected_suggestion = list.get(list.curselection()[0])
            selected_suggestion = selected_suggestion.split()
            Key = selected_suggestion[0]

            mark_index_entry.delete(0, END)
            mark_index_entry.insert(0, Key)

            list.place_forget()

        mark_index_entry = Entry(form_6_9_edit, fg="#1b2d52", font=fonts)
        mark_index_entry.place(width=203, height=26, x=270, y=183)
        mark_index_entry.bind("<KeyRelease>", lambda event: collect(mark_index_entry.get()))

        Year = Entry(form_6_9_edit, fg="#1b2d52", font=fonts)
        Year.place(width=150, height=25, x=532, y=183)
        Year.insert(0, "None")
        Year.configure(state="disabled")

        options = ["1st", "2nd", "3rd"]
        Term = ttk.Combobox(form_6_9_edit, values=options, font=fonts)
        Term.place(width=150, height=25, x=736, y=183)
        Term.set(options[0])

        options_g6 = ["6A", "6B", "6C", "6D", "6E", "6F"]
        options_g7 = ["7A", "7B", "7C", "7D", "7E", "7F"]
        options_g8 = ["8A", "8B", "8C", "8D", "8E", "8F"]
        options_g9 = ["9A", "9B", "9C", "9D", "9E", "9F"]
        options = options_g6 + options_g7 + options_g8 + options_g9
        Class = ttk.Combobox(form_6_9_edit, values=options, font=fonts)
        Class.place(width=150, height=25, x=938, y=183)
        Class.set(options[0])
    except:
        pass

    # Sheet
    try:
        sinhala_mark_entry = Entry(form_6_9_edit, fg="#1b2d52", font=fonts, bg="#e9edfa")
        sinhala_mark_entry.place(width=78, height=30, x=370, y=256)
        maths_mark_entry = Entry(form_6_9_edit, fg="#1b2d52", font=fonts, bg="#e9edfa")
        maths_mark_entry.place(width=78, height=30, x=595, y=256)
        science_mark_entry = Entry(form_6_9_edit, fg="#1b2d52", font=fonts, bg="#e9edfa")
        science_mark_entry.place(width=78, height=30, x=834, y=256)
        english_mark_entry = Entry(form_6_9_edit, fg="#1b2d52", font=fonts, bg="#e9edfa")
        english_mark_entry.place(width=78, height=30, x=1057, y=256)
        religion_mark_entry = Entry(form_6_9_edit, fg="#1b2d52", font=fonts, bg="#e9edfa")
        religion_mark_entry.place(width=78, height=30, x=370, y=291)
        history_mark_entry = Entry(form_6_9_edit, fg="#1b2d52", font=fonts, bg="#e9edfa")
        history_mark_entry.place(width=78, height=30, x=595, y=291)
        geography_mark_entry = Entry(form_6_9_edit, fg="#1b2d52", font=fonts, bg="#e9edfa")
        geography_mark_entry.place(width=78, height=30, x=834, y=291)
        ce_ls_mark_entry = Entry(form_6_9_edit, fg="#1b2d52", font=fonts, bg="#e9edfa")
        ce_ls_mark_entry.place(width=78, height=30, x=1057, y=291)
        health_mark_entry = Entry(form_6_9_edit, fg="#1b2d52", font=fonts, bg="#e9edfa")
        health_mark_entry.place(width=78, height=30, x=370, y=325)
        tamil_mark_entry = Entry(form_6_9_edit, fg="#1b2d52", font=fonts, bg="#e9edfa")
        tamil_mark_entry.place(width=78, height=30, x=595, y=325)
        pts_mark_entry = Entry(form_6_9_edit, fg="#1b2d52", font=fonts, bg="#e9edfa")
        pts_mark_entry.place(width=78, height=30, x=834, y=325)
        ict_mark_entry = Entry(form_6_9_edit, fg="#1b2d52", font=fonts, bg="#e9edfa")
        ict_mark_entry.place(width=78, height=30, x=1057, y=325)

        art_mark_entry = Entry(form_6_9_edit, fg="#1b2d52", font=fonts, bg="#e9edfa")
        art_mark_entry.place(width=78, height=30, x=370, y=370)
        dance_mark_entry = Entry(form_6_9_edit, fg="#1b2d52", font=fonts, bg="#e9edfa")
        dance_mark_entry.place(width=78, height=30, x=595, y=370)
        music_mark_entry = Entry(form_6_9_edit, fg="#1b2d52", font=fonts, bg="#e9edfa")
        music_mark_entry.place(width=78, height=30, x=834, y=370)
        drama_mark_entry = Entry(form_6_9_edit, fg="#1b2d52", font=fonts, bg="#e9edfa")
        drama_mark_entry.place(width=78, height=30, x=1057, y=370)
    except:
        pass

    fonts = ("Candara", 14, "bold")

    def filler():
        if art_mark_entry.get() == "":
            art_mark_entry.insert(0, "#")
        if drama_mark_entry.get() == "":
            drama_mark_entry.insert(0, "#")
        if dance_mark_entry.get() == "":
            dance_mark_entry.insert(0, "#")
        if music_mark_entry.get() == "":
            music_mark_entry.insert(0, "#")
        if ict_mark_entry.get() == "":
            ict_mark_entry.insert(0, "#")
        save()

    def search():
        if mark_index_entry.get() == "" or \
                Year.get() == "" or \
                Term.get() == "" or \
                Class.get() == "":
            messagebox.showerror("Error", "Index , Year, Term, Class Required")
        else:
            try:
                key = winreg.OpenKey(winreg.HKEY_CURRENT_USER, r"SOFTWARE\PADMA")
                host = winreg.QueryValueEx(key, "Host")[0]
                user = winreg.QueryValueEx(key, "User")[0]
                password = winreg.QueryValueEx(key, "Pass")[0]
                winreg.CloseKey(key)

                con = mysql.connector.connect(host=host,
                                              user=user,
                                              password=password,
                                              database="pcc")
                mycursor = con.cursor(buffered=True)
            except:
                messagebox.showerror("Error", "Database Failed")
                return

            try:
                query = "create table exam(id integer,year integer,term varchar(5),class varchar(10)" \
                        ",exam_id integer PRIMARY KEY AUTO_INCREMENT)"
                mycursor.execute(query)
                query = "create table mark_uo(exam_id integer PRIMARY KEY AUTO_INCREMENT" \
                        ",sinhala integer" \
                        ",maths integer" \
                        ",science integer" \
                        ",english integer" \
                        ",religion integer" \
                        ",history integer" \
                        ",geograpy integer" \
                        ",ce integer" \
                        ",health integer" \
                        ",art varchar(5)" \
                        ",dance varchar(5)" \
                        ",music varchar(5)" \
                        ",drama varchar(5)" \
                        ",tamil integer" \
                        ",pts integer" \
                        ",ict integer)"
                mycursor.execute(query)
            except:
                mycursor.execute("use pcc")

            query = 'select * from student where id=%s'
            parameters = [mark_index_entry.get()]
            mycursor.execute(query, parameters)
            result = mycursor.fetchone()

            if result is not None:
                query = "select * from exam where id=%s and term=%s and class=%s"
                parameters = [mark_index_entry.get(), Term.get(), Class.get()]
                mycursor.execute(query, parameters)
                result = mycursor.fetchone()

                if result is not None:
                    query = """
                            SELECT student.id, mark_uo.*, student.first_name, student.last_name
                            FROM exam
                            INNER JOIN mark_uo ON exam.exam_id = mark_uo.exam_id
                            INNER JOIN student ON exam.id = student.id 
                            where exam.id=%s and exam.term=%s and exam.class=%s 
                            """
                    parameters = [mark_index_entry.get(), Term.get(), Class.get()]
                    mycursor.execute(query, parameters)
                    result = mycursor.fetchone()
                    con.close()

                    if result is not None:
                        sinhala_mark_entry.delete(0, END)
                        maths_mark_entry.delete(0, END)
                        science_mark_entry.delete(0, END)
                        english_mark_entry.delete(0, END)
                        religion_mark_entry.delete(0, END)
                        history_mark_entry.delete(0, END)
                        geography_mark_entry.delete(0, END)
                        ce_ls_mark_entry.delete(0, END)
                        health_mark_entry.delete(0, END)
                        art_mark_entry.delete(0, END)
                        dance_mark_entry.delete(0, END)
                        music_mark_entry.delete(0, END)
                        drama_mark_entry.delete(0, END)
                        tamil_mark_entry.delete(0, END)
                        pts_mark_entry.delete(0, END)
                        ict_mark_entry.delete(0, END)

                        sinhala_mark_entry.insert(0, result[2])
                        maths_mark_entry.insert(0, result[3])
                        science_mark_entry.insert(0, result[4])
                        english_mark_entry.insert(0, result[5])
                        religion_mark_entry.insert(0, result[6])
                        history_mark_entry.insert(0, result[7])
                        geography_mark_entry.insert(0, result[8])
                        ce_ls_mark_entry.insert(0, result[9])
                        health_mark_entry.insert(0, result[10])
                        art_mark_entry.insert(0, result[11])
                        dance_mark_entry.insert(0, result[12])
                        music_mark_entry.insert(0, result[13])
                        drama_mark_entry.insert(0, result[14])
                        tamil_mark_entry.insert(0, result[15])
                        pts_mark_entry.insert(0, result[16])
                        ict_mark_entry.insert(0, result[17])
                    else:
                        pass
                else:
                    messagebox.showerror("Error", "No Result Found")
            else:
                messagebox.showerror("Error", "Invalid index or Unknown Student")
            con.close()
            print("pass")

    def save():
        if mark_index_entry.get() == "" or \
                Year.get() == "" or \
                Term.get() == "" or \
                Class.get() == "":
            messagebox.showerror("Error", "Index , Year, Term, Class Required")

        elif religion_mark_entry.get() == "" or \
                sinhala_mark_entry.get() == "" or \
                english_mark_entry.get() == "" or \
                tamil_mark_entry.get() == "" or \
                science_mark_entry.get() == "" or \
                maths_mark_entry.get() == "" or \
                history_mark_entry.get() == "" or \
                health_mark_entry.get() == "" or \
                pts_mark_entry.get() == "" or \
                geography_mark_entry.get() == "" or \
                ce_ls_mark_entry.get() == "":
            messagebox.showerror("Error", "All Marks Required")
        elif art_mark_entry.get() == "" or dance_mark_entry.get() == "" or drama_mark_entry.get() == "" or \
                music_mark_entry.get() == "" or ict_mark_entry.get() == "":
            filler()
        else:
            try:
                key = winreg.OpenKey(winreg.HKEY_CURRENT_USER, r"SOFTWARE\PADMA")
                host = winreg.QueryValueEx(key, "Host")[0]
                user = winreg.QueryValueEx(key, "User")[0]
                password = winreg.QueryValueEx(key, "Pass")[0]
                winreg.CloseKey(key)

                con = mysql.connector.connect(host=host,
                                              user=user,
                                              password=password,
                                              database="pcc")
                mycursor = con.cursor(buffered=True)
            except:
                messagebox.showerror("Error", "Database Failed")
                return

            try:
                query = "create table exam(id integer,year integer,term varchar(5),class varchar(10)" \
                        ",exam_id integer PRIMARY KEY AUTO_INCREMENT)"
                mycursor.execute(query)
                query = "create table mark_uo(exam_id integer PRIMARY KEY AUTO_INCREMENT" \
                        ",sinhala integer" \
                        ",maths integer" \
                        ",science integer" \
                        ",english integer" \
                        ",religion integer" \
                        ",history integer" \
                        ",geograpy integer" \
                        ",ce integer" \
                        ",health integer" \
                        ",art varchar(5)" \
                        ",dance varchar(5)" \
                        ",music varchar(5)" \
                        ",drama varchar(5)" \
                        ",tamil integer" \
                        ",pts integer" \
                        ",ict varchar(5))"
                mycursor.execute(query)
            except:
                mycursor.execute("use pcc")

            query = "UPDATE mark_uo " \
                    "INNER JOIN exam ON mark_uo.exam_id = exam.exam_id " \
                    "SET mark_uo.sinhala = %s, mark_uo.maths = %s, mark_uo.science = %s, mark_uo.english = %s, " \
                    "mark_uo.religion = %s, mark_uo.history = %s, mark_uo.geograpy = %s, mark_uo.ce = %s, " \
                    "mark_uo.health = %s, " \
                    "mark_uo.art = %s, mark_uo.dance = %s, mark_uo.music = %s, mark_uo.drama = %s, mark_uo.tamil = " \
                    "%s, mark_uo.pts = %s, mark_uo.ict = %s " \
                    "where exam.id=%s and exam.term=%s and exam.class=%s"

            parameters = [
                sinhala_mark_entry.get(),
                maths_mark_entry.get(),
                science_mark_entry.get(),
                english_mark_entry.get(),
                religion_mark_entry.get(),
                history_mark_entry.get(),
                geography_mark_entry.get(),
                ce_ls_mark_entry.get(),
                health_mark_entry.get(),
                art_mark_entry.get(),
                dance_mark_entry.get(),
                music_mark_entry.get(),
                drama_mark_entry.get(),
                tamil_mark_entry.get(),
                pts_mark_entry.get(),
                ict_mark_entry.get(),
                mark_index_entry.get(),
                Term.get(),
                Class.get()
            ]

            mycursor.execute(query, parameters)
            con.commit()
            con.close()
            messagebox.showinfo("Success", "DATA UPDATED SUCCESSFULLY")

    list = Listbox(form_6_9_edit, height=10, width=35, bd=0)
    list.bind("<Double-Button-1>", lambda event: select())

    def show():
        if list.winfo_ismapped():
            list.place_forget()
        else:
            list.place(x=270, y=209)

    drop_btn = Button(form_6_9_edit, text="⇅", bg="white", font=("Candara", 20, "bold"), bd=0, command=show)
    drop_btn.place(x=445, y=184, width=26, height=25)

    edit_marks_SEARCH = Button(form_6_9_edit, text="SEARCH", bg="#ce4912", fg="white", font=fonts, command=search)
    edit_marks_SEARCH.place(x=253, y=416, width=172, height=34)

    edit_marks_SAVE = Button(form_6_9_edit, text="SAVE", bg="#3b3086", fg="white", font=fonts, command=save)
    edit_marks_SAVE.place(x=445, y=416, width=172, height=34)

    EDIT = Button(form_6_9_edit, text="ADD MARK", bg="#3b3086", fg="white", font=fonts
                  , command=lambda: (form_6_9_edit.destroy(), Mark_form_6_9(frame)))
    EDIT.place(x=22, y=295, width=172, height=34)

    VIEWER = Button(form_6_9_edit, text="VIEWER", bg="#ce4912", fg="white", font=fonts
                    , command=lambda: (form_6_9_edit.destroy(), Viewer(frame)))
    VIEWER.place(x=22, y=340, width=172, height=34)


def Viewer(frame):
    viewer = Frame(frame, height=651, width=1185)
    viewer.pack()

    back = PhotoImage(file=current+"/.ux/padma/.mark/.viewer/viewer.png")
    Background = Label(viewer, image=back)
    Background.image = back
    Background.place(x=0, y=0, relheight=1, relwidth=1)

    global x
    x = date.today()
    x = x.year

    def find():
        for item in view.get_children():
            view.delete(item)

        if radio_var.get() == 0:
            print("Student Mode")
            if mark_index_entry.get() == "" or mark_class_combo.get() == "" or term.get() == "":
                messagebox.showerror("Error", "Index, Term, Class are Required !")
            else:
                try:
                    key = winreg.OpenKey(winreg.HKEY_CURRENT_USER, r"SOFTWARE\PADMA")
                    host = winreg.QueryValueEx(key, "Host")[0]
                    user = winreg.QueryValueEx(key, "User")[0]
                    password = winreg.QueryValueEx(key, "Pass")[0]
                    winreg.CloseKey(key)

                    con = mysql.connector.connect(host=host,
                                                  user=user,
                                                  password=password,
                                                  database="pcc")
                    mycursor = con.cursor(buffered=True)
                except:
                    messagebox.showerror("Error", "Database Fail")
                    return

                query = "select * from exam where id=%s"
                parameters = [mark_index_entry.get()]
                mycursor.execute(query, parameters)
                result = mycursor.fetchone()

                if result is not None:
                    query = """
                           SELECT student.id, mark_uo.*, student.first_name, student.last_name
                           FROM exam
                           INNER JOIN mark_uo ON exam.exam_id = mark_uo.exam_id
                           INNER JOIN student ON exam.id = student.id 
                           where exam.id=%s and exam.term=%s and exam.class=%s 
                           """
                    parameters = [mark_index_entry.get(), term.get(), mark_class_combo.get()]
                    mycursor.execute(query, parameters)
                    result = mycursor.fetchall()
                    con.close()

                    counter = 0
                    for row in result:
                        counter += 1
                        view.insert(parent="", index=1, iid=counter, text="", values=(
                            row[0], row[18] + " " + row[19], row[2], row[3], row[4], row[5], row[6], row[7], row[8],
                            row[9],
                            row[10], row[11], row[12], row[13], row[14], row[15], row[16], row[17]
                        ))
                else:
                    messagebox.showerror("Error", "Invalid Index !")
        elif radio_var.get() == 1:
            print("Class Mode")
            if mark_year_entry.get() == "" or mark_class_combo.get() == "" or term.get() == "":
                messagebox.showerror("Error", "Year, Term, Class are Required !")
            else:
                try:
                    key = winreg.OpenKey(winreg.HKEY_CURRENT_USER, r"SOFTWARE\PADMA")
                    host = winreg.QueryValueEx(key, "Host")[0]
                    user = winreg.QueryValueEx(key, "User")[0]
                    password = winreg.QueryValueEx(key, "Pass")[0]
                    winreg.CloseKey(key)

                    con = mysql.connector.connect(host=host,
                                                  user=user,
                                                  password=password,
                                                  database="pcc")
                    mycursor = con.cursor()
                except:
                    messagebox.showerror("Error", "Database Fail")
                    return

                query = """
                       SELECT student.id, mark_uo.*, student.first_name, student.last_name
                       FROM exam
                       INNER JOIN mark_uo ON exam.exam_id = mark_uo.exam_id
                       INNER JOIN student ON exam.id = student.id 
                       where exam.year=%s and exam.term=%s and exam.class=%s 
                       """
                parameters = [mark_year_entry.get(), term.get(), mark_class_combo.get()]
                mycursor.execute(query, parameters)
                result = mycursor.fetchall()
                con.close()

                counter = 0
                for row in result:
                    counter += 1
                    view.insert(parent="", index=1, iid=counter, text="", values=(
                        row[0], row[18] + " " + row[19], row[2], row[3], row[4], row[5], row[6], row[7], row[8], row[9],
                        row[10], row[11], row[12], row[13], row[14], row[15], row[16], row[17]
                    ))

    fonts = ("Candara", 15)

    # Header
    try:
        mark_index_entry = Entry(viewer, fg="#1b2d52", font=fonts)
        mark_index_entry.place(width=203, height=26, x=57, y=149)

        global mark_year_entry
        mark_year_entry = Entry(viewer, fg="#1b2d52", font=fonts)
        mark_year_entry.place(width=150, height=26, x=279, y=149)
        mark_year_entry.insert(0, x)

        global term
        options = ["1st", "2nd", "3rd"]
        term = ttk.Combobox(viewer, values=options, font=fonts)
        term.place(width=150, height=26, x=449, y=149)
        term.set(options[0])

        global mark_class_combo
        options_g6 = ["6A", "6B", "6C", "6D", "6E", "6F"]
        options_g7 = ["7A", "7B", "7C", "7D", "7E", "7F"]
        options_g8 = ["8A", "6B", "8C", "8D", "8E", "8F"]
        options_g9 = ["9A", "9B", "9C", "9D", "9E", "9F"]
        options = options_g6 + options_g7 + options_g8 + options_g9
        mark_class_combo = ttk.Combobox(viewer, values=options, font=fonts)
        mark_class_combo.place(width=150, height=26, x=618, y=149)
        mark_class_combo.set(options[0])

        find_marks_button = Button(viewer, text="🔎", bg="#3b3086", fg="white", font=fonts, command=find)
        find_marks_button.place(x=780, y=149, width=66, height=30)

        global excel_class_combo
        options_eg = ["6", "7", "8", "9"]
        excel_class_combo = ttk.Combobox(viewer, values=options_eg, font=fonts)
        excel_class_combo.place(width=79, height=26, x=883, y=149)
        excel_class_combo.set(options_eg[0])

        global excel_sheet_combo
        options_sheet = ["A", "B", "C", "D", "E", "F", "G"]
        excel_sheet_combo = ttk.Combobox(viewer, values=options_sheet, font=fonts)
        excel_sheet_combo.place(width=79, height=26, x=989, y=149)
        excel_sheet_combo.set(options_sheet[0])

        def move():
            if radio_var.get() == 0:
                messagebox.showerror("Error", "you should activate the \"class mode\" before transfer data to excel")
            elif radio_var.get() == 1:
                if mark_year_entry.get() == "" or term.get() == "" or mark_class_combo.get() == "" or \
                        excel_class_combo.get() == "" or excel_sheet_combo.get() == "":
                    messagebox.showerror("error", "Please enter all details without Index !")
                else:
                    Transfer(excel_sheet_combo.get())
            else:
                messagebox.showerror("error", "please activate class mode or student mode")

        move_marks_button = Button(viewer, text="↔", bg="#3b3086", fg="white", font=fonts, command=move)
        move_marks_button.place(x=1080, y=149, width=53, height=30)

        mark_year_entry.configure(state="disabled")

        def ss():
            select_class.configure(value=1)
            mark_year_entry.configure(state="disabled")
            mark_index_entry.configure(state="normal")

        radio_var = IntVar()
        select_student = tk.Radiobutton(viewer, value=0, variable=radio_var, text="Student Mode", bg="#d2daf4"
                                        , anchor=W)
        select_student.place(x=1041, y=31, height=15, width=100)
        select_student.configure(bg="#c9d2f1", activebackground="#c9d2f1")
        select_student.configure(command=ss)

        def sc():
            mark_index_entry.configure(state="disabled")
            mark_year_entry.configure(state="normal")
            mark_class_combo.configure(state="normal")
            select_class.configure(value=1)

        select_class = tk.Radiobutton(viewer, value=1, variable=radio_var, text="Class Mode", bg="#d2daf4", anchor=W)
        select_class.place(x=1041, y=51, height=15, width=100)
        select_class.configure(bg="#c9d2f1", activebackground="#c9d2f1")
        select_class.configure(command=sc)
    except:
        pass

    try:
        view = ttk.Treeview(viewer)
        view.place(x=22, y=216, height=422, width=1142)

        view["columns"] = (
            "1", "2", "3", "4", "5", "6", "7", "8", "9",
            "10", "11", "12", "13", "14", "15", "16", "17", "18"
        )

        view["show"] = "headings"

        view.column("1", minwidth=10, width=15)
        view.column("2", minwidth=20, width=100)
        view.column("3", minwidth=5, width=10)
        view.column("4", minwidth=5, width=10)
        view.column("5", minwidth=5, width=10)
        view.column("6", minwidth=5, width=10)
        view.column("7", minwidth=5, width=10)
        view.column("8", minwidth=5, width=10)
        view.column("9", minwidth=5, width=10)
        view.column("10", minwidth=5, width=10)
        view.column("11", minwidth=5, width=10)
        view.column("12", minwidth=5, width=10)
        view.column("13", minwidth=5, width=10)
        view.column("14", minwidth=5, width=10)
        view.column("15", minwidth=5, width=10)
        view.column("16", minwidth=5, width=10)
        view.column("17", minwidth=5, width=10)
        view.column("18", minwidth=5, width=10)

        view.heading("1", text="Index")
        view.heading("2", text="Name")
        view.heading("3", text="Sinhala")
        view.heading("4", text="Math")
        view.heading("5", text="Science")
        view.heading("6", text="English")
        view.heading("7", text="Religion")
        view.heading("8", text="History")
        view.heading("9", text="Geo")
        view.heading("10", text="C.E")
        view.heading("11", text="Health")
        view.heading("12", text="Art")
        view.heading("13", text="Dance")
        view.heading("14", text="Music")
        view.heading("15", text="Drama")
        view.heading("16", text="Tamil")
        view.heading("17", text="P.T.S")
        view.heading("18", text="I.C.T")
    except:
        pass


def Transfer(wsheet):
    print("Class Mode")
    if mark_year_entry.get() == "" or mark_class_combo.get() == "" or term.get() == "":
        messagebox.showerror("Error", "Year, Term, Class are Required !")
    else:
        try:
            key = winreg.OpenKey(winreg.HKEY_CURRENT_USER, r"SOFTWARE\PADMA")
            host = winreg.QueryValueEx(key, "Host")[0]
            user = winreg.QueryValueEx(key, "User")[0]
            password = winreg.QueryValueEx(key, "Pass")[0]
            winreg.CloseKey(key)

            con = mysql.connector.connect(host=host,
                                          user=user,
                                          password=password,
                                          database="pcc")
            mycursor = con.cursor()
        except:
            messagebox.showerror("Error", "Database Fail")
            return

        query = """
               SELECT student.id, mark_uo.*, student.first_name, student.last_name
               FROM exam
               INNER JOIN mark_uo ON exam.exam_id = mark_uo.exam_id
               INNER JOIN student ON exam.id = student.id 
               where exam.year=%s and exam.term=%s and exam.class=%s 
               """
        parameters = [mark_year_entry.get(), term.get(), mark_class_combo.get()]
        mycursor.execute(query, parameters)
        result = mycursor.fetchall()
        con.close()

        open_location = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx;*.xls")])
        # excel_file_path = "E:/Py Soft/Padmawathie/Temp.xlsx"
        # Load the workbook
        if open_location != "":
            workbook = openpyxl.load_workbook(open_location)

            # Select a specific worksheet by name
            sheet = workbook[wsheet]  # Replace "A" with the name of your desired worksheet

            sheet[f"B{1}"] = None
            sheet[f"B{1}"] = excel_class_combo.get() + excel_sheet_combo.get()

            row_index = 10
            for row in result:
                sheet[f"D{row_index}"] = None  # Column1
                sheet[f"E{row_index}"] = None  # Column2

                sheet[f"F{row_index}"] = None  # Column3
                sheet[f"H{row_index}"] = None  # Column4
                sheet[f"J{row_index}"] = None  # Column5
                sheet[f"L{row_index}"] = None  # Column6
                sheet[f"N{row_index}"] = None  # Column7
                sheet[f"P{row_index}"] = None  # Column8
                sheet[f"R{row_index}"] = None  # Column9
                sheet[f"T{row_index}"] = None  # Column10
                sheet[f"V{row_index}"] = None  # Column11
                sheet[f"X{row_index}"] = None  # Column12
                sheet[f"Z{row_index}"] = None  # Column13
                sheet[f"AB{row_index}"] = None  # Column14
                sheet[f"AD{row_index}"] = None  # Column15
                sheet[f"AF{row_index}"] = None  # Column16
                sheet[f"AH{row_index}"] = None  # Column17
                sheet[f"AJ{row_index}"] = None  # Column18

                sheet[f"D{row_index}"] = row[0]  # Column1
                sheet[f"E{row_index}"] = row[18] + " " + row[19]  # Column2

                sheet[f"F{row_index}"] = row[2]  # Column3
                sheet[f"H{row_index}"] = row[3]  # Column4
                sheet[f"J{row_index}"] = row[4]  # Column5
                sheet[f"L{row_index}"] = row[5]  # Column6
                sheet[f"N{row_index}"] = row[6]  # Column7
                sheet[f"P{row_index}"] = row[7]  # Column8
                sheet[f"R{row_index}"] = row[8]  # Column9
                sheet[f"T{row_index}"] = row[9]  # Column10
                sheet[f"V{row_index}"] = row[10]  # Column11
                sheet[f"X{row_index}"] = row[11]  # Column12
                sheet[f"Z{row_index}"] = row[12]  # Column13
                sheet[f"AB{row_index}"] = row[13]  # Column14
                sheet[f"AD{row_index}"] = row[14]  # Column15
                sheet[f"AF{row_index}"] = row[15]  # Column16
                sheet[f"AH{row_index}"] = row[16]  # Column17
                sheet[f"AJ{row_index}"] = row[17]  # Column18
                row_index += 1

            # Save the Excel file
            save_location = filedialog.asksaveasfilename(
                title="Select a file",
                filetypes=[("Excel", "*.xlsx"), ("JPEG", "*.jpg"), ("PNG", "*.png"), ("PDF", "*.pdf")],
                defaultextension=".xlsx"
            )
            if save_location == "":
                pass
            else:
                workbook.save(save_location)
        else:
            pass


def Report(frame):
    Reporte = Frame(frame, height=651, width=1185)
    Reporte.pack()

    back = PhotoImage(file=current+"/.ux/padma/.mark/.report/Report_mark.png")
    Background = Label(Reporte, image=back)
    Background.image = back
    Background.place(x=0, y=0, relheight=1, relwidth=1)

    global x
    x = date.today()
    x = x.year

    fonts = ("Candara", 15)

    try:
        mark_index_entry = Entry(Reporte, fg="#1b2d52", font=fonts)
        mark_index_entry.place(width=203, height=26, x=57, y=149)

        global mark_year_entry
        mark_year_entry = Entry(Reporte, fg="#1b2d52", font=fonts)
        mark_year_entry.place(width=150, height=26, x=279, y=149)
        mark_year_entry.insert(0, x)

        global term
        options = ["1st", "2nd", "3rd"]
        term = ttk.Combobox(Reporte, values=options, font=fonts)
        term.place(width=150, height=26, x=449, y=149)
        term.set(options[0])

        global mark_class_combo
        options_g6 = ["6A", "6B", "6C", "6D", "6E", "6F"]
        options_g7 = ["7A", "7B", "7C", "7D", "7E", "7F"]
        options_g8 = ["8A", "6B", "8C", "8D", "8E", "8F"]
        options_g9 = ["9A", "9B", "9C", "9D", "9E", "9F"]
        options = options_g6 + options_g7 + options_g8 + options_g9
        mark_class_combo = ttk.Combobox(Reporte, values=options, font=fonts)
        mark_class_combo.place(width=150, height=26, x=618, y=149)
        mark_class_combo.set(options[0])

        find_marks_button = Button(Reporte, text="/", bg="#3b3086", fg="white", font=fonts, state="disabled")
        find_marks_button.place(x=780, y=149, width=66, height=30)

        global excel_class_combo
        options_eg = ["6", "7", "8", "9"]
        excel_class_combo = ttk.Combobox(Reporte, values=options_eg, font=fonts, state="disabled")
        excel_class_combo.place(width=79, height=26, x=883, y=149)
        excel_class_combo.set(options_eg[0])

        global excel_sheet_combo
        options_sheet = ["A", "B", "C", "D", "E", "F", "G"]
        excel_sheet_combo = ttk.Combobox(Reporte, values=options_sheet, font=fonts, state="disabled")
        excel_sheet_combo.place(width=79, height=26, x=989, y=149)
        excel_sheet_combo.set(options_sheet[0])

        move_marks_button = Button(Reporte, text="*", bg="#3b3086", fg="white", font=fonts, state="disabled")
        move_marks_button.place(x=1080, y=149, width=53, height=30)

        mark_year_entry.configure(state="disabled")

        def ss():
            select_class.configure(value=1)
            mark_year_entry.configure(state="disabled")
            mark_index_entry.configure(state="normal")

        radio_var = IntVar()
        select_student = tk.Radiobutton(Reporte, value=0, variable=radio_var, text="Student Mode", bg="#d2daf4"
                                        , anchor=W)
        select_student.place(x=1041, y=31, height=15, width=100)
        select_student.configure(bg="#c9d2f1", activebackground="#c9d2f1")
        select_student.configure(command=ss)

        def sc():
            mark_index_entry.configure(state="disabled")
            mark_year_entry.configure(state="normal")
            mark_class_combo.configure(state="normal")
            select_class.configure(value=1)

        select_class = tk.Radiobutton(Reporte, value=1, variable=radio_var, text="Class Mode", bg="#d2daf4", anchor=W)
        select_class.place(x=1041, y=51, height=15, width=100)
        select_class.configure(bg="#c9d2f1", activebackground="#c9d2f1")
        select_class.configure(command=sc)
    except:
        pass

    def student_report():
        if radio_var.get() == 1:
            messagebox.showerror("Error", "You're in Class Mode, Please select student mode to get student mark sheet")
        elif radio_var.get() == 0:
            try:
                key = winreg.OpenKey(winreg.HKEY_CURRENT_USER, r"SOFTWARE\PADMA")
                host = winreg.QueryValueEx(key, "Host")[0]
                user = winreg.QueryValueEx(key, "User")[0]
                password = winreg.QueryValueEx(key, "Pass")[0]
                winreg.CloseKey(key)

                conn = mysql.connector.connect(host=host,
                                               user=user,
                                               password=password,
                                               database="pcc")
                mycursor = conn.cursor(buffered=True)
                # print("connection ok !")
            except:
                messagebox.showerror("Error", "DataBase Failed")
                return

            query = "select * from exam where id=%s"
            parameters = [mark_index_entry.get()]
            mycursor.execute(query, parameters)
            result = mycursor.fetchone()

            if result is not None:
                query = """
                       SELECT student.id, mark_uo.*, student.first_name, student.last_name
                       FROM exam
                       INNER JOIN mark_uo ON exam.exam_id = mark_uo.exam_id
                       INNER JOIN student ON exam.id = student.id 
                       where exam.id=%s and exam.term=%s and exam.class=%s 
                       """
                parameters = [mark_index_entry.get(), term.get(), mark_class_combo.get()]
                mycursor.execute(query, parameters)
                results = mycursor.fetchall()

                pdf = FPDF(orientation='L', unit='mm')
                pdf.set_auto_page_break(auto=False, margin=0)
                pdf.set_font("Arial", size=30, style='B')
                pdf.add_page()

                title_x = 20
                title_y = 10
                pdf.set_xy(title_x, title_y)
                pdf.cell(0, 10,
                         "Padmavathie N.C - " + mark_index_entry.get() + ", " + mark_class_combo.get() + " - Mark Sheet"
                         , ln=True, align='L')

                subtitle_x = 50
                subtitle_y = title_y + 15
                pdf.set_xy(subtitle_x, subtitle_y)
                pdf.set_font("Arial", size=0, style='I')
                pdf.cell(0, 10, "", ln=True, align='L')

                # Set font and font size for table content
                pdf.set_font("Arial", size=8)
                # Calculate column widths
                header = [i[0] for i in mycursor.description]  # Get column names from mycursor
                column_widths = [pdf.get_string_width(str(header[i])) for i in range(len(header))]

                # Iterate over the results and update column widths
                for row in results:
                    for i in range(len(row)):
                        width = pdf.get_string_width(str(row[i]))
                        if width > column_widths[i]:
                            column_widths[i] = width

                # Calculate total table width
                table_width = sum(column_widths)

                # Calculate starting position to fit the table from side to side
                start_x = (pdf.w - table_width) / 12

                # Set column widths and display the table
                pdf.set_x(start_x)
                margin = 2  # Define the margin value

                # Add the table headings
                for header_name, width in zip(header, column_widths):
                    pdf.cell(width + 2 * margin, 10, str(header_name), border=1, ln=False,
                             align='C')  # Adjusted cell width

                pdf.ln()

                for row in results:
                    # Check if adding a new row will exceed the page height
                    if pdf.get_y() + 10 > pdf.h - 10:
                        pdf.add_page()  # Add a new page
                        pdf.set_xy(start_x, 10)  # Set position for the table on the new page

                        # Add the column headers
                        for header_name, width in zip(header, column_widths):
                            pdf.cell(width + 2 * margin, 10, str(header_name), border=1, ln=False,
                                     align='C')  # Adjusted cell width

                        pdf.ln()

                    # Display the row
                    pdf.set_x(start_x)
                    for i in range(len(row)):
                        pdf.cell(column_widths[i] + 2 * margin, 10, str(row[i]), border=1, ln=False,
                                 align='C')  # Adjusted cell width
                    pdf.ln()

                save_location = filedialog.asksaveasfilename(
                    title="Select a file",
                    filetypes=[("JPEG", "*.jpg"), ("PNG", "*.png"), ("PDF", "*.pdf")],
                    defaultextension=".pdf"
                )
                if save_location == "":
                    pass
                else:
                    pdf.output(save_location)

                mycursor.close()
                conn.close()
            else:
                messagebox.showerror("Error", "Invalid Student Id Or No Data Found")
        else:
            messagebox.showerror("Error", "Mode Selector Error!")

    def class_report():
        if radio_var.get() == 0:
            messagebox.showerror("Error"
                                 , "You're in Student Mode, Please select Class mode to get student mark sheet")
        elif radio_var.get() == 1:
            try:
                key = winreg.OpenKey(winreg.HKEY_CURRENT_USER, r"SOFTWARE\PADMA")
                host = winreg.QueryValueEx(key, "Host")[0]
                user = winreg.QueryValueEx(key, "User")[0]
                password = winreg.QueryValueEx(key, "Pass")[0]
                winreg.CloseKey(key)

                conn = mysql.connector.connect(host=host,
                                               user=user,
                                               password=password,
                                               database="pcc")
                mycursor = conn.cursor(buffered=True)
                # print("connection ok !")
            except:
                messagebox.showerror("Error", "DataBase Failed")
                return

            query = "select * from exam where id=%s"
            parameters = [mark_index_entry.get()]
            mycursor.execute(query, parameters)
            print("test1 pass")

            query = """
                                   SELECT student.id, mark_uo.*, student.first_name, student.last_name
                                   FROM exam
                                   INNER JOIN mark_uo ON exam.exam_id = mark_uo.exam_id
                                   INNER JOIN student ON exam.id = student.id 
                                   where exam.year=%s and exam.term=%s and exam.class=%s 
                                   """
            parameters = [mark_year_entry.get(), term.get(), mark_class_combo.get()]
            mycursor.execute(query, parameters)
            results = mycursor.fetchall()
            print("test2 pass")

            pdf = FPDF(orientation='L', unit='mm')
            pdf.set_auto_page_break(auto=False, margin=0)
            pdf.set_font("Arial", size=30, style='B')
            pdf.add_page()
            print("test3 pass")

            title_x = 20
            title_y = 10
            pdf.set_xy(title_x, title_y)
            pdf.cell(0, 10,
                     "Padmavathie N.C - " + mark_year_entry.get() + ", " + mark_class_combo.get() + " - Mark Sheet"
                     , ln=True, align='L')

            subtitle_x = 50
            subtitle_y = title_y + 15
            pdf.set_xy(subtitle_x, subtitle_y)
            pdf.set_font("Arial", size=0, style='I')
            pdf.cell(0, 10, "", ln=True, align='L')
            print("test4 pass")

            # Set font and font size for table content
            pdf.set_font("Arial", size=8)
            # Calculate column widths
            header = [i[0] for i in mycursor.description]  # Get column names from mycursor
            column_widths = [pdf.get_string_width(str(header[i])) for i in range(len(header))]
            print("test1 pass")

            # Iterate over the results and update column widths
            for row in results:
                for i in range(len(row)):
                    width = pdf.get_string_width(str(row[i]))
                    if width > column_widths[i]:
                        column_widths[i] = width
            print("test1 pass")

            # Calculate total table width
            table_width = sum(column_widths)

            # Calculate starting position to fit the table from side to side
            start_x = (pdf.w - table_width) / 12

            # Set column widths and display the table
            pdf.set_x(start_x)
            margin = 2  # Define the margin value

            # Add the table headings
            for header_name, width in zip(header, column_widths):
                pdf.cell(width + 2 * margin, 10, str(header_name), border=1, ln=False,
                         align='C')  # Adjusted cell width

            pdf.ln()
            print("test1 pass")

            for row in results:
                # Check if adding a new row will exceed the page height
                if pdf.get_y() + 10 > pdf.h - 10:
                    pdf.add_page()  # Add a new page
                    pdf.set_xy(start_x, 10)  # Set position for the table on the new page

                    # Add the column headers
                    for header_name, width in zip(header, column_widths):
                        pdf.cell(width + 2 * margin, 10, str(header_name), border=1, ln=False,
                                 align='C')  # Adjusted cell width

                    pdf.ln()
                print("test1 pass")

                # Display the row
                pdf.set_x(start_x)
                for i in range(len(row)):
                    pdf.cell(column_widths[i] + 2 * margin, 10, str(row[i]), border=1, ln=False,
                             align='C')  # Adjusted cell width
                pdf.ln()

            save_location = filedialog.asksaveasfilename(
                title="Select a file",
                filetypes=[("JPEG", "*.jpg"), ("PNG", "*.png"), ("PDF", "*.pdf")],
                defaultextension=".pdf"
            )
            if save_location == "":
                pass
            else:
                pdf.output(save_location)
                print("test1 pass")

            mycursor.close()
            conn.close()
        else:
            messagebox.showerror("Error", "Mode Selector Error!")
            print("test1 pass")

    st_btn = Button(Reporte, bg="#3b3086", fg="white", text="Generate", font=fonts, command=student_report)
    st_btn.place(x=171, y=514, width=159, height=38)

    class_btn = Button(Reporte, bg="#ce4912", fg="white", text="Generate", font=fonts, command=class_report)
    class_btn.place(x=847, y=514, width=159, height=38)
